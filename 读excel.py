import openpyxl

#wb = openpyxl.load_workbook('demo.xlsx')
#sheet = wb["Sheet1"]
# print(sheet["B1"].value)
# print(sheet["I1"].value)
# print(sheet["O1"].value)
# print(sheet["U1"].value)
# # 红字 + 黑色填充
#cell = sheet["C13"]
wb1 = openpyxl.load_workbook('a.xlsx')
sheet1 = wb1.active
# print(sheet["B1"].value)
# print(sheet["I1"].value)
# print(sheet["O1"].value)
# print(sheet["U1"].value)
# # 红字 + 黑色填充
cell1 = sheet1["H12"]
#print('1111111', cell.font)
# print(cell.value, cell.font.color.rgb[2:], cell.fill, cell.has_style, cell.style, dir(cell))

# print('22222222', cell1.font)
# print(cell.value, cell.font.color.rgb[2:], cell.fill, cell.has_style, cell.style, dir(cell))
#print('2222222', cell.fill)
# red = openpyxl.styles.colors.Color(rgb='FF00B0F0', indexed=None, auto=None, theme=None,
#                                   tint = 0.0, type = 'rgb')
# print(red)
# print('2222222', cell1.fill)
# 定义颜色
blue = openpyxl.styles.colors.Color(rgb='FF00B0F0', type='rgb')
black = openpyxl.styles.colors.Color(rgb=None, theme=1, type='theme')
red = openpyxl.styles.colors.Color(rgb='FFFF0000', type='rgb')
white = openpyxl.styles.colors.Color(rgb=None, theme=0, type='theme')
# 字体
blue_font = openpyxl.styles.Font(name='等线', charset=134, family=2.0,
                                       b=False, i=False, color=blue, sz=12.0, scheme='minor')
red_font = openpyxl.styles.Font(name='等线', charset=134, family=2.0, b=False,
                                      i=False, color=red, sz=12.0, scheme='minor')
black_font = openpyxl.styles.Font(name='等线', charset=134, family=2.0,
                                        b=False, i=False, color=black, sz=12.0, scheme='minor')
white_font = openpyxl.styles.Font(name='等线', charset=134, family=2.0,
                                        b=False, i=False, color=white, sz=12.0, scheme='minor')
# 填充
red_fill = openpyxl.styles.PatternFill("solid", fgColor=red)
black_fill = openpyxl.styles.PatternFill("solid", fgColor=black)
other_fill = openpyxl.styles.PatternFill()
cell1.font=blue_font
cell1.fill=black_fill
cell1.value='测试字体'
wb1.save('c.xlsx')
